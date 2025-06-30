import requests 
import openpyxl
RED = '\033[91m'
GREEN = '\033[92m'
YELLOW = '\033[93m'
BLUE = '\033[94m'
RESET = '\033[0m'
dic_lat = {}
dic_lon = {}
# functions:
# def search(Name):
    
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         dic_lat.setdefault(row[1], row[2])
#         dic_lon.setdefault(row[1], row[3])

print(GREEN + "__________Welcom To Weather App__________" + RESET)     
api_key = "41f4d894f7e41dd9783d34b0a55f47bc"
exel_address = "F:\\python_prj\\myself.pr\\weather app\\cities.xlsx"
workbook = openpyxl.load_workbook(exel_address)
sheet = workbook.active  
choose = ""
while choose != 0:
    print(GREEN + "Pls Choice a City from bottom list and write name in persian or english:")
    i = 1
    print("""
1. Show list
0.Exit"""+ RESET)
    choose = int(input(RED +"Choice a number: " + RESET))
    if choose == 1:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            dic_lat.setdefault(row[1], row[2])
            dic_lon.setdefault(row[1], row[3])
            print(BLUE + f'{i}', f") {row[1]}" + RESET)
            i += 1
        city_name = str(input(RED + "Enter name of That city You want: " + RESET))
        url = f"https://api.openweathermap.org/data/2.5/weather?lat={dic_lat[city_name]}&lon={dic_lon[city_name]}&appid={api_key}"

        response = requests.get(url)
        if response.status_code== 200:
            data = response.json()

            print(YELLOW + f"City: {data['name']}")
            print(f"Temperature: {data['main']['temp']} Kelvin")
            selisuse = round(data['main']['temp'] - 273, 2)
            print(f"Temperature: {selisuse} Celsius")
            print(f"Pressure : {data['main']['pressure']} pa")
            print(f"Humidity : %{data['main']['humidity']} ")
            print(f"Minimum Temperature : {round(data['main']['temp_min']-273 , 2)} Celsius")
            print(f"Maximum Temperature: {round(data['main']['temp_max'] - 273, 2)} Celsius")
            print(f"Weather condition : {data['weather'][0]['description']}" + RESET)
            

        else:
            print(f"خطا: {response.status_code}")